#### pakcage needed
import sys
import os
import platform
import re
import pdfplumber
import pandas as pd
from PyQt5.QtWidgets import QApplication, QMainWindow, QLabel, QVBoxLayout, QPushButton, QWidget, QFileDialog, QScrollArea, QGridLayout, QCheckBox, QSizePolicy, QSpacerItem , QHBoxLayout, QMessageBox
from PyQt5.QtGui import QPixmap, QImage, QPainter, QFont
from PyQt5.QtCore import Qt, QSize, QPoint, QFileInfo, pyqtSignal
from PIL import Image
from fpdf import FPDF
from openpyxl import load_workbook
import json
from functools import partial
#######

### FUNCTIONS
# Extracting data, saving images, and PDFs remain unchanged 
def extract_from_pdf(invoice_path):
    """
    Extracts model codes and quantities from a PDF invoice.

    Opens the PDF at `invoice_path`, searching each line for "NR" and extracting 
    model codes with both letters and numbers, along with their quantities.

    Args:
        invoice_path (str): Path to the PDF file.

    Returns:
        list of tuple: List of (model code, quantity) pairs as strings.
    """
    models_quantities = []
    code_pattern = re.compile(r'^(?=.*[A-Z])(?=.*[0-9])[A-Z0-9]+$')  # Pattern to match codes with both letters and numbers

    with pdfplumber.open(invoice_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            lines = text.split('\n')

            for line in lines:
                if "NR" in line:
                    columns = line.split()
                    try:
                        nr_index = columns.index("NR")
                        model = columns[0]
                        quantity = columns[nr_index + 1]
                        if code_pattern.match(model):
                            models_quantities.append((model, quantity))
                    except (ValueError, IndexError):
                        continue
    return models_quantities

## Extracting data, saving images from an xlsx file
def extract_from_xlsx(invoice_path):
    """
    Extracts model codes and quantities from an Excel invoice.

    Args:
        invoice_path (str): Path to the Excel (.xlsx or .xlsm) file.

    Returns:
        list of list: A list of [model, quantity] pairs, or an empty list if no valid data is found.
    """
    models_quantities = []
    try:
        # Load the workbook and get the first sheet
        wb = load_workbook(invoice_path, data_only=True)  # data_only=True to get computed values
        sheet = wb.active

        # Unmerge merged cells (iterate over a copy of the ranges)
        for merged_cell_range in list(sheet.merged_cells.ranges):
            sheet.unmerge_cells(str(merged_cell_range))

        # Propagate merged cell values in the 'Quantity' column (assumed to be the 12th column in the sheet)
        for merged_cell_range in list(sheet.merged_cells.ranges):
            min_col, min_row, max_col, max_row = merged_cell_range.bounds
            if min_col == 12:  # This assumes the 'Quantity' column is the 12th column (index 12)
                value = sheet.cell(row=min_row, column=min_col).value
                for row in range(min_row + 1, max_row + 1):
                    sheet.cell(row=row, column=min_col).value = value

        # Load data into pandas after unmerging cells
        data = sheet.values
        df = pd.DataFrame(data)

        # Use row 5 (index 4) as the header and limit rows to 6-34
        df.columns = df.iloc[4].fillna(method="ffill")  # Fill merged header cells
        df = df[5:35].reset_index(drop=True)  # Rows 6 to 34 (inclusive)

        # Remove columns that are duplicates or irrelevant
        df = df.loc[:, ~df.columns.duplicated()]  # Remove duplicate columns
        df = df[["Model", "Quantity"]]  # Keep only the relevant columns

        # Ensure "Item" and "Quantity" columns exist
        if "Model" in df.columns and "Quantity" in df.columns:
            # Strip spaces and convert to numeric
            df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce")  # Convert to numeric, NaN for errors
            df[["Model", "Quantity"]] = df[["Model", "Quantity"]].applymap(
                lambda x: str(x).strip() if isinstance(x, str) else x
            )
            
            # Filter rows where "Item" is valid and "Quantity" is not NaN
            filtered_df = df[["Model", "Quantity"]]
            filtered_df = filtered_df[
                (filtered_df["Model"].notna() & (filtered_df["Model"] != "")) &
                (filtered_df["Quantity"].notna())  # Quantity must not be NaN
            ]

            # Debug: Show filtered data
            #print("Filtered data (before conversion):")
            #print(filtered_df)

            # Convert to a list of lists
            models_quantities = filtered_df.values.tolist()
        else:
            print("Required columns 'Model' and 'Quantity' not found.")
    except Exception as e:
        print(f"Failed to read XLSX file: {e}")
    
    return models_quantities

#Checkboxes
class CustomCheckBox(QLabel):
    """
    A custom checkbox class represented by an image, toggling between checked and unchecked states.

    This QLabel-based checkbox displays an image that changes on click, simulating a 
    checkbox toggle. The checkbox is initially unchecked.

    Attributes:
        is_checked (bool): Tracks the checkbox's state; True if checked, False otherwise.

    Methods:
        mousePressEvent(event): Toggles the checkbox state and updates the displayed image.
    """
    stateChanged = pyqtSignal(bool)
    def __init__(self):
        super().__init__()
        
        # Get the base path for resources
        box_base_path = os.path.dirname(__file__)

        # Define paths to the checked and unchecked images
        self.unchecked_path = os.path.join(box_base_path, "supp_img", "unchecked.png")
        self.checked_path = os.path.join(box_base_path, "supp_img", "checked.png")

        # Set initial properties for the checkbox
        self.setFixedSize(30, 30)
        self.setPixmap(QPixmap(self.unchecked_path).scaled(30, 30, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        self.is_checked = False

    # Function to tick the checkboxes
    def mousePressEvent(self, event):
        # Switch between checked and unchecked images
        if self.is_checked:
            self.setPixmap(QPixmap(self.unchecked_path).scaled(30, 30, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        else:
            self.setPixmap(QPixmap(self.checked_path).scaled(30, 30, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        self.is_checked = not self.is_checked
        self.stateChanged.emit(self.is_checked)

class InvoiceSelectionWindow(QMainWindow):
    """
    A main window for selecting and displaying an invoice, with logo, description, button, and footer.

    This window allows the user to select a PDF or Excel file representing an invoice. It extracts
    model codes and quantities and opens an image display window with the extracted data. The layout 
    includes a logo, description, a button to select an invoice, and a footer message.

    Attributes:
        logo_path (str): Path to the logo image file.
        invoice_title (str): Title extracted from the selected invoice file name.

    Methods:
        resizeEvent(event): Updates the logo image size on window resize.
        update_logo_pixmap(): Scales and sets the logo image.
        on_submit(): Opens a file dialog to select an invoice and extracts its data.
        extract_invoice_title(file_path): Extracts the file name (without extension) as the invoice title.
        open_image_display_window(models_quantities): Opens a new window to display the selected invoice data.
    """
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Invoice Choice")
        self.setMinimumSize(600, 400)  # Set the fixed size for the window
        # Set the background color to white
        self.setStyleSheet("background-color: white;")

        self.layout = QVBoxLayout()
    
        # LOGO
        self.logo_path = os.path.join(os.path.dirname(__file__), "supp_img", "logo.jpg" ) # Update this to your logo path
        self.logo_label = QLabel(self)
        self.logo_label.setAlignment(Qt.AlignCenter)  # Center the logo in the layout
        self.layout.addWidget(self.logo_label)

        # DESCRIPTION
        self.description_label = QLabel("Select the invoice to display products' picture, code, and quantity")
        self.description_label.setAlignment(Qt.AlignCenter)  # Center the description text
        font = QFont("Arial", 17)  # Change font family and size
        self.description_label.setFont(font)
        self.layout.addWidget(self.description_label)

        # BUTTON
        self.button = QPushButton("Select Invoice", self) #what's written inside the button
        self.button.clicked.connect(self.on_submit) #what it does when it is clicked
        self.button.setFont(QFont("Arial", 18)) #font of the button
        self.button.setFixedWidth(200) # fixed width for the button (change it to make it smaller/bigger)
        # Style the button
        self.button.setStyleSheet("""
            QPushButton {
                background-color: #8C000F;  /* crimson background */
                border: 2px solid #650021;   /* Darker crimson border */
                border-radius: 15px;          /* Rounded corners */
                color: white;                 /* White text */
                padding: 10px 20px;           /* Padding around text */
                font-size: 16px;              /* Text size */
            }
            QPushButton:hover {
                background-color: #650021;    /* Darker crimson on hover */
            }
        """)
        # Add a spacer above the button to control the spacing between the description and the button
        self.layout.addSpacing(20)  # Small space between description and button
        # Center the button using an HBox layout
        button_layout = QHBoxLayout()
        button_layout.addStretch(1)  # Add stretchable space before the button
        button_layout.addWidget(self.button)
        button_layout.addStretch(1)  # Add stretchable space after the button
        self.layout.addLayout(button_layout)

        #FOOTER
        # Spacer above the footer to control its position
        self.layout.addSpacerItem(QSpacerItem(20, 20, QSizePolicy.Minimum, QSizePolicy.Expanding))
        # Footer writing
        self.footer_label = QLabel("2024. All rights reserved", self)
        self.footer_label.setAlignment(Qt.AlignCenter)  # Center the footer text
        footer_font = QFont("Arial", 10)  # Change font family and size for the footer
        self.footer_label.setFont(footer_font)
        self.layout.addWidget(self.footer_label)


        container = QWidget()
        container.setLayout(self.layout)
        self.setCentralWidget(container)

        # Set the initial pixmap
        self.update_logo_pixmap()

    def resizeEvent(self, event):
        """Update the logo pixmap on window resize."""
        self.update_logo_pixmap()
        super().resizeEvent(event)

    def update_logo_pixmap(self):
        """Update the pixmap to scale with the window size."""
        pixmap = QPixmap(self.logo_path)
        if pixmap.isNull():
            print("Error: Unable to load logo image.")
            return
        scaled_pixmap = pixmap.scaled(self.width(), self.height() // 2, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        self.logo_label.setPixmap(scaled_pixmap)
        
    # select the invoice and obtain all the info needed
    def on_submit(self):
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(self, "Select Invoice", "", "PDF Files (*.pdf);;Excel Files (*.xlsx *.xlsm)", options=options)
        
        if file_path:
            # Extract the invoice title from the file path
            self.invoice_title = self.extract_invoice_title(file_path)
            
            # Extract data based on file type (PDF or Excel)
            if file_path.endswith(".pdf"):
                models_quantities = extract_from_pdf(file_path)
            elif file_path.endswith((".xlsx", ".xlsm")):
                models_quantities = extract_from_xlsx(file_path)
            else:
                return

            # 1. Create and open the ImageDisplayWindow
            self.open_image_display_window(models_quantities)

            # 2. After opening, load the checkbox states with a delay (if needed)
            # Using QTimer to ensure the checkbox states are loaded after the window is shown
            from PyQt5.QtCore import QTimer

            # Delay the load of checkbox states to ensure the window is fully initialized
            QTimer.singleShot(0, lambda: self.image_display_window.load_checkbox_states())

    def extract_invoice_title(self, file_path):
        """Extract the title from the invoice file path."""
        return QFileInfo(file_path).baseName() # Get the file name without the path and extension

    def open_image_display_window(self, models_quantities):
        self.image_display_window = ImageDisplayWindow(models_quantities, self.invoice_title)
        self.image_display_window.show()

class ImageDisplayWindow(QMainWindow):
    """
    A window for displaying images, model codes, quantities, and checkboxes 
    in a 3x3 grid layout, with the option to save the display as a PDF.

    Attributes:
        invoice_title (str): The title of the invoice displayed in the window and PDF.
        layout (QVBoxLayout): Main layout that contains the scroll area and button layout.
        image_paths (list): Stores paths to images associated with each model.
        scroll_area (QScrollArea): Scroll area for navigating through product images.
        grid_layout (QGridLayout): Layout for arranging images, text, and checkboxes in a grid.
        save_pdf_button (QPushButton): Button that triggers saving the display as a PDF.
        widgets (list): References to all dynamically created widgets in the grid layout.

    Methods:
        __init__(models_quantities, invoice_title):
            Initializes the window, setting the layout, widgets, and styles.
        display_models_and_quantities(models_quantities):
            Displays the provided models and quantities in the 3x3 grid layout.
        resizeEvent(event):
            Overrides the resize event to adjust widget sizes dynamically.
        resize_widgets():
            Adjusts the sizes of images, text, and checkboxes based on the window size.
        resize_cells_to_fit_page(page_width_px, page_height_px):
            Resizes cells and content to fit an A4 page, used for PDF generation.
        save_as_pdf():
            Saves the current display to a multi-page PDF, formatted with images, 
            model codes, quantities, and checkboxes.
    """
    def __init__(self, models_quantities, invoice_title):
        super().__init__()
        self.invoice_title = invoice_title #unique identifier for the invoice
        self.setWindowTitle(f"{self.invoice_title}")
        self.setGeometry(100, 100, 1200, 900)
        self.checkbox_states = {} #hold the invoice checkboxes
        
        self.widgets = [] #initilize widgets list early
    
        # Initialize layout and widgets
        self.layout = QVBoxLayout()
        self.layout.setSpacing(0) #set spacing between elements to 0
        self.layout.setContentsMargins(2, 2, 2, 2)
        self.image_paths = []
        
        #TITLE
        # Create a layout
        #title_layout = QVBoxLayout()
        #self.title_label = QLabel(f"{self.invoice_title}")  # Title label
        #self.title_label.setAlignment(Qt.AlignCenter)  # Center the title
        #self.title_label.setStyleSheet("font-size: 16px; font-weight: bold;")  # Style the title
        #title_layout.addWidget(self.title_label)  # Add title label to the title layout
        #self.layout.addLayout(title_layout)

        #SCROLL AREA
        # Scroll Area for product images
        self.scroll_area = QScrollArea()
        self.scroll_area_widget = QWidget()
        self.scroll_area.setWidget(self.scroll_area_widget)
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setContentsMargins(0, 0, 0, 0)
        # Grid Layout for 3x3 display
        self.grid_layout = QGridLayout()
        self.scroll_area_widget.setLayout(self.grid_layout)
        self.layout.addWidget(self.scroll_area)

        #BUTTON
        # Create a horizontal layout to center the Save PDF button
        pdfbutton_layout = QHBoxLayout()
        pdfbutton_layout.setContentsMargins(0, 10, 0, 0)  # Set margins to zero
        pdfbutton_layout.setSpacing(0)  # Set spacing to 0 between items in the layout
        # Save PDF Button
        self.save_pdf_button = QPushButton("Save as PDF", self)  # What's written inside the button
        self.save_pdf_button.clicked.connect(self.save_as_pdf)  # What it does when clicked
        self.save_pdf_button.setFixedWidth(200)  # Fixed width for the button
        self.save_pdf_button.setFixedHeight(30)  # Fixed height for the button (changeable)
        # Add stretchable space before and after the button to center it
        pdfbutton_layout.addStretch(1)  # Stretchable space before the button
        pdfbutton_layout.addWidget(self.save_pdf_button)  # Add the button to the layout
        pdfbutton_layout.addStretch(1)  # Stretchable space after the button
        # Add the button layout to the main layout
        self.layout.addLayout(pdfbutton_layout)
        # Add a spacer item to reduce the space taken by the button layout
        spacer_item = QSpacerItem(20, 10, QSizePolicy.Minimum, QSizePolicy.Fixed)  # Adjust height as needed
        self.layout.addItem(spacer_item)

        # Container for the main layout
        container = QWidget()
        container.setLayout(self.layout)
        self.setCentralWidget(container)

        # Set stylesheet for the entire window and components
        self.setStyleSheet("""
            QMainWindow {
                background-color: white;
            }
            QWidget {
                background-color: white;
            }
            QScrollBar:vertical {
                background: white; /* Background of the scrollbar area */
                width: 7px; /* Scrollbar width */
                margin: 0px 0px 0px 0px; /* margin for the scrollbar */
            }
            QScrollBar::handle:vertical {
                background: #C0C0C0; /* Scrollbar handle color */
                border-radius: 3px; /* Rounded corners for the handle */
                min-height: 10px; /* Minimum height of the handle */
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                background: none; /* No background for the add and sub lines */
            }
            QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
                background: #f0f0f0; /* backgorund for the scroll area outside tha hande */
            }
            QPushButton {
                background-color: #8C000F;  /* crimson background */
                color: white;
                border: 2px solid #650021;   /* Darker crimson border */
                border-radius: 15px; /* Rounded corners */
                padding: 3px;
                font-size: 12px;  /* Ensure button text is visible */
            }
            QPushButton:hover {
                background-color: #650021;
            }
            QPushButton:pressed {
                background-color: #555;
            }
        """)

        # Display images, code, quantity, and checkbox
        self.display_models_and_quantities(models_quantities)
        
        # load saved checkbox states
        self.load_checkbox_states()
       
    def show_order_ready_notification(self):
        """Displays a customized notification when the order is ready."""
        # Create a QMessageBox instance
        msg_box = QMessageBox(self)
        msg_box.setIcon(QMessageBox.Information)  # Set an information icon
        msg_box.setWindowTitle("Order Complete")  # Set the title
        msg_box.setText("✅ The order is ready!")  # Set the main message
        msg_box.setInformativeText("📦 Order is now ready for shipping! 📦")  # Add additional information

        # Add buttons and style them
        ok_button = msg_box.addButton(QMessageBox.Ok)
        ok_button.setText("OK")  # Customize the button text
        ok_button.setStyleSheet("font-weight: bold; color: #f9f9f9; background-color: #8C000F; border-radius: 10px; padding: 5px;")
        
        # Set styles for the message box
        msg_box.setStyleSheet("""
            QMessageBox {
                background-color: #f9f9f9;  /* Main background color */
                border: none;  /* Remove default border */
                border-radius: 10px;  /* Rounded corners */
                padding: 0;  /* Remove extra padding */
                margin: 0;  /* Remove extra margin */
            }
            QMessageBox QLabel {
                font-size: 14px;  /* Main text font size */
                color: #000;  /* Text color */
                background-color: #f9f9f9;  /* Match QLabel background to QMessageBox */
                border: none;  /* Remove any QLabel borders */
                margin: 0;
                padding: 0;
            }
            QMessageBox QLabel#qt_msgboxex_icon_label {
                background-color: #f9f9f9;  /* Match QMessageBox background */
                border: none;  /* Remove border */
            }
            QMessageBox QPushButton {
                background-color: #f0f0f0;  /* Button background */
                color: #000;  /* Button text color */
                font-weight: bold;  /* Bold button text */
                border: 1px solid #8C000F;  /* Button border color */
                border-radius: 5px;  /* Rounded corners for buttons */
                padding: 5px 10px;  /* Button padding */
                margin: 5px;  /* Space between buttons */
            }
            QMessageBox QPushButton:hover {
                background-color: #e8e8e8;  /* Lighter background on hover */
            }
        """)

        # Show the customized message box
        msg_box.exec_()
    
    def check_order_ready(self):
            all_checked = all(checkbox.is_checked for _, _, _, checkbox, _, _ in self.widgets)
            if all_checked:
                self.show_order_ready_notification()
                
    def save_checkbox_states(self):
        """Save the current checkbox states to a JSON file, linked to the invoice."""
        self.json_path = os.path.join(os.path.dirname(__file__), "checkbox_states.json")
        try:
            # Load existing data from the file
            with open(self.json_path, "r") as file:
                data = json.load(file)
        except FileNotFoundError:
            data = {}

        # Update or add the current invoice's checkbox states
        data[self.invoice_title] = {}
        for img_label, model_label, quantity_label, checkbox, model, quantity in self.widgets:
            data[self.invoice_title][model] = checkbox.is_checked

        # Save the updated data back to the file
        with open(self.json_path, "w") as file:
            json.dump(data, file)
            
        print(f"Saved checkbox states: {data}")
        self.check_order_ready()

    def load_checkbox_states(self):
        """Load the checkbox states from the JSON file."""
        self.json_path = os.path.join(os.path.dirname(__file__), "checkbox_states.json")
        try:
            # Load saved states from the file
            with open(self.json_path, "r") as file:
                self.checkbox_states = json.load(file)
        except FileNotFoundError:
            print("No checkbox_states.json file found. Skipping loading.")
            self.checkbox_states = {}
            return
        except json.JSONDecodeError:
            print("Error decoding checkbox_states.json. Skipping loading.")
            self.checkbox_states = {}
            return
        # Check if states exist for the current invoice
        if self.invoice_title not in self.checkbox_states:
            print(f"No saved checkbox states available for invoice: {self.invoice_title}")
            return
        # Retrieve states for the current invoice
        saved_states = self.checkbox_states[self.invoice_title]
        # Define paths for checked and unchecked images
        box_base_path = os.path.dirname(__file__)
        unchecked_path = os.path.join(box_base_path, "supp_img", "unchecked.png")
        checked_path = os.path.join(box_base_path, "supp_img", "checked.png")
        # Apply saved states to widgets
        for _, _, _, checkbox, model, _ in self.widgets:
            is_checked = saved_states.get(model, False)  # Default to unchecked
            checkbox.is_checked = is_checked
            if is_checked:
                checkbox.setPixmap(QPixmap(checked_path).scaled(30, 30, Qt.KeepAspectRatio, Qt.SmoothTransformation))
            else:
                checkbox.setPixmap(QPixmap(unchecked_path).scaled(30, 30, Qt.KeepAspectRatio, Qt.SmoothTransformation))
            print(f"Loaded state for model {model}: {'Checked' if is_checked else 'Unchecked'}")

    def closeEvent(self, event):
        """Override the close event to save the checkbox states for the current invoice."""
        self.save_checkbox_states()
        event.accept()

    def display_models_and_quantities(self, models_quantities):
        placeholder_image_path = os.path.join(os.path.dirname(__file__), "supp_img", "img_not_found.jpg")
        self.widgets = []  # Store references to dynamically resized widgets

        for _, _, _, checkbox, model, _ in self.widgets:
            #set the paths
            box_base_path = os.path.dirname(__file__)
            checkbox.unchecked_path = os.path.join(box_base_path, "supp_img", "unchecked.png")
            checkbox.checked_path = os.path.join(box_base_path, "supp_img", "checked.png")
            # Set the checkbox state from the dictionary
            checkbox.is_checked = self.checkbox_states.get(model, False)
            
            # Update the checkbox appearance manually based on the state
            if checkbox.is_checked:
                checkbox.setPixmap(QPixmap(checkbox.checked_path).scaled(30, 30, Qt.KeepAspectRatio, Qt.SmoothTransformation))
            else:
                checkbox.setPixmap(QPixmap(checkbox.unchecked_path).scaled(30, 30, Qt.KeepAspectRatio, Qt.SmoothTransformation))
            
        # Clear the grid layout first
        while self.grid_layout.count():
            child = self.grid_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

        for idx, (model, quantity) in enumerate(models_quantities):
            images_folder_path = os.path.join(os.path.dirname(__file__), "product_img", f"{model}.jpg")
            # Create widgets
            img_label = QLabel()
            model_label = QLabel(f"Model: {model}")
            quantity_label = QLabel(f"Quantity: {quantity}")
            checkbox = CustomCheckBox()
            
            # Set initial properties
            img_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            model_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            quantity_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

            # Center the image and text inside their cells
            img_label.setAlignment(Qt.AlignCenter)
            model_label.setAlignment(Qt.AlignCenter)
            quantity_label.setAlignment(Qt.AlignCenter)

            # Create a vertical layout for model and quantity to display them one under the other
            text_layout = QVBoxLayout()
            text_layout.addWidget(model_label)
            text_layout.addWidget(quantity_label)
            text_layout.setSpacing(5)

            # Create a widget to hold the layout
            text_widget = QWidget()
            text_widget.setLayout(text_layout)

            # Style the checkbox (center it and make it larger)
            checkbox.setStyleSheet("""
                QCheckBox::indicator {
                    width: 35px;
                    height: 35px;
                }
            """)
            checkbox.setText("")  # Remove text from the checkbox

            # Center the checkbox within the cell using a layout
            checkbox_layout = QVBoxLayout()
            checkbox_layout.addWidget(checkbox)
            checkbox_layout.setAlignment(Qt.AlignCenter)

            checkbox_widget = QWidget()
            checkbox_widget.setLayout(checkbox_layout)
            # Connect signal
            checkbox.stateChanged.connect(partial(self.update_checkbox_state, model))

            # Initialize state in the dictionary
            checkbox.is_checked = self.checkbox_states.get(model, False)

            # Connect stateChanged signal to update the dictionary
            checkbox.stateChanged.connect(lambda state, model=model: self.update_checkbox_state(model, state))

            # Add widgets to the grid layout
            row = (idx // 3) * 3
            col = idx % 3
            self.grid_layout.addWidget(img_label, row, col)
            self.grid_layout.addWidget(text_widget, row + 1, col)
            self.grid_layout.addWidget(checkbox_widget, row + 2, col)

            # Store widget references
            self.widgets.append((img_label, model_label, quantity_label, checkbox, model, quantity))

        # Add spacer items to fill empty space
        for row in range((len(models_quantities) // 3) * 3 + 3, 9):
            for col in range(3):
                spacer = QSpacerItem(1, 1, QSizePolicy.Expanding, QSizePolicy.Expanding)
                self.grid_layout.addItem(spacer, row, col)

        # Initial resize
        self.resize_widgets()
        
    def update_checkbox_state(self, model, state):
        """
        Updates the state of a checkbox for a given model in the checkbox_states dictionary.
        """
        self.checkbox_states[model] = state
        print(f"Checkbox for {model} updated to {'checked' if state else 'unchecked'}")

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.resize_widgets()
        
    def resize_widgets(self):
        # Get the current size of the window
        window_size = self.size()
        cell_width = int(window_size.width() / 3)
        cell_height = int(window_size.height() / 3)

        for img_label, model_label, quantity_label, checkbox, model, quantity in self.widgets:
            # Adjust image size
            pixmap = QPixmap()
            image_path = os.path.join(os.path.dirname(__file__), "product_img", f"{model}.jpg")
            if os.path.exists(image_path):
                pixmap = QPixmap(image_path)
            
                # Apply scaling factor (e.g., 1.1 to increase the size by 10%)
                scaling_factor = 1.2
                pixmap = pixmap.scaled(int(cell_width * scaling_factor), int(cell_height * scaling_factor),
                                   Qt.KeepAspectRatio, Qt.SmoothTransformation)
                img_label.setPixmap(pixmap)
            else:
                imgNotfoundPath = os.path.join(os.path.dirname(__file__), "supp_img", "img_not_found.jpg")
                pixmap = QPixmap(imgNotfoundPath)  # Load placeholder image
                                # Apply scaling factor (e.g., 1.1 to increase the size by 10%)
                scaling_factor = 1.2
                pixmap = pixmap.scaled(int(cell_width * scaling_factor), int(cell_height * scaling_factor),
                                   Qt.KeepAspectRatio, Qt.SmoothTransformation)
                img_label.setPixmap(pixmap)

            # Adjust text size (scaling is approximate) and center the text
            font_size = int(cell_height / 20)
            model_label.setStyleSheet(f"font-size: {font_size}px;")
            quantity_label.setStyleSheet(f"font-size: {font_size}px;")
            model_label.setAlignment(Qt.AlignCenter)
            quantity_label.setAlignment(Qt.AlignCenter)

            # Adjust checkbox size
            checkbox.setStyleSheet(f"QCheckBox::indicator {{ width: {int(cell_width / 8)}px; height: {int(cell_width / 8)}px; }}")

    def resize_cells_to_fit_page(self, page_width_px, page_height_px):
        # This method should resize the cells and their content to fit the A4 page size
        cell_width = page_width_px / 3
        cell_height = page_height_px / 3

        for img_label, model_label, quantity_label, checkbox, model, quantity in self.widgets:
            # Adjust image size
            pixmap = QPixmap()
            image_path = os.path.join(os.path.dirname(__file__), "product_img", f"{model}.jpg")
            if os.path.exists(image_path):
                pixmap = QPixmap(image_path)
            else:
                # Load the placeholder image instead
                imgNotfoundPath = os.path.join(os.path.dirname(__file__), "supp_img", "img_not_found.jpg")
                pixmap = QPixmap(imgNotfoundPath)

            # Apply scaling factor to ensure the image fits the cell
            pixmap = pixmap.scaled(int(cell_width), int(cell_height),
                                Qt.KeepAspectRatio, Qt.SmoothTransformation)
            img_label.setPixmap(pixmap)

            # Adjust text size
            font_size = int(cell_height / 20)
            model_label.setStyleSheet(f"font-size: {font_size}px;")
            quantity_label.setStyleSheet(f"font-size: {font_size}px;")
            model_label.setAlignment(Qt.AlignCenter)
            quantity_label.setAlignment(Qt.AlignCenter)

            # Adjust checkbox size
            checkbox.setStyleSheet(f"QCheckBox::indicator {{ width: {int(cell_width / 8)}px; height: {int(cell_width / 8)}px; }}")
    
    def save_as_pdf(self):
        # Create a PDF document
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=10)
        
        # Define the margins
        top_margin = 1  # Minimal top margin to fit the title without affecting the grid
        bottom_margin = 10  # Bottom margin in mm

        # Define A4 dimensions in mm
        pdf_width = 210  # A4 width
        pdf_height = 297  # A4 height
        
        items_per_page = 9
        total_items = len(self.widgets)
        pages_needed = (total_items // items_per_page) + (1 if total_items % items_per_page > 0 else 0)

        for page in range(pages_needed):
            pdf.add_page()
            
            # Add title to the top of the page
            pdf.set_font("Arial", 'B', size=8)  # Small font for the title
            pdf.set_xy(5, top_margin)  # Position the title as close as possible to the top
            pdf.cell(pdf_width - 10, 5, f"{self.invoice_title}", 0, 1, 'C')  # Centered title, with minimal height

            # Reset font for the grid items
            pdf.set_font("Arial", size=8)

            for i in range(items_per_page):
                item_index = page * items_per_page + i
                if item_index >= total_items:
                    break

                img_label, model_label, quantity_label, checkbox, model, quantity = self.widgets[item_index]

                # Set image path
                images_folder_path = os.path.join(os.path.dirname(__file__), "product_img")
                image_path = os.path.join(images_folder_path, f"{model}.jpg")

                if not os.path.exists(image_path):
                    image_path = os.path.join(os.path.dirname(__file__), "supp_img", "img_not_found.jpg") #Placeholder image path

                # Image dimensions
                img_width = 53
                img_height = 72  # Keep aspect ratio in mind if necessary
                
                # Calculate position for the 3x3 grid
                col = i % 3
                row = i // 3
                x_position = col * (pdf_width / 3)
                y_position = top_margin + 6 + row * (img_height + 20)  # Adjusted for minimal title height

                # Place the image in the PDF
                pdf.image(image_path, x=x_position + (pdf_width / 3 - img_width) / 2, 
                        y=y_position, w=img_width, h=img_height)

                # Set a smaller font size for the text
                pdf.set_font("Arial", size=10)  # Font size for text

                # Model and Quantity Text
                pdf.set_xy(x_position, y_position + img_height + 1)  # Position for Model text
                pdf.cell(pdf_width / 3, 5, f"Model: {model}", ln=True, align='C')  # Center text within column width

                # Set Y position for Quantity text just below Model text
                pdf.set_xy(x_position, y_position + img_height + 6)  # Position for Quantity text
                pdf.cell(pdf_width / 3, 5, f"Quantity: {quantity}", ln=True, align='C')  # Center text within column width
                
                # Checkbox Image
                # Checkbox Image
                checkbox_image_path = os.path.join(
                    os.path.dirname(__file__), 
                    "supp_img", 
                    "checked.png" if checkbox.is_checked else "unchecked.png"
                )
                checkbox_offset = 3
                checkbox_x = x_position + (pdf_width / 3 - 10) / 2 + checkbox_offset  # Center checkbox
                checkbox_y = y_position + img_height + 12  # Below the quantity
                pdf.image(checkbox_image_path, x=checkbox_x, y=checkbox_y, w=6, h=6)
            
                
                # Add bottom margin
                pdf.ln(bottom_margin / 10)

        # Save the PDF to file
        output_pdf_path = os.path.join(os.path.expanduser("~/Desktop"), f"{self.invoice_title}_parts.pdf")
        pdf.output(output_pdf_path)
        print(f"Saved PDF path: '{output_pdf_path}'")

        # Automatically open the PDF file after saving 
        self.open_pdf_file(output_pdf_path)
  
    def open_pdf_file(self, pdf_path):
        """Automatically open the PDF based on the operating system."""
        if platform.system() == "Darwin":  # macOS
            os.system(f"open {pdf_path}")
        elif platform.system() == "Windows":  # Windows
            os.startfile(pdf_path)
        elif platform.system() == "Linux":  # Linux
            os.system(f"xdg-open {pdf_path}")
        
    def create_invoice_grid(self):
        # Clear previous grid layout
        for i in reversed(range(self.grid_layout.count())):
            widget = self.grid_layout.itemAt(i).widget()
            if widget is not None:
                widget.deleteLater()

        # Set the number of rows and columns
        num_columns = 3
        num_rows = (len(self.image_labels) + num_columns - 1) // num_columns

        # Add images, codes, quantities, and checkboxes to the grid
        for idx, image_label in enumerate(self.image_labels):
            row = idx // num_columns
            col = idx % num_columns

            # Create a vertical layout for each cell
            cell_layout = QVBoxLayout()

            # Add image
            if image_label.image is not None:
                image_item = QLabel()
                image_item.setPixmap(image_label.image.scaled(100, 100, Qt.KeepAspectRatio, Qt.SmoothTransformation))
                cell_layout.addWidget(image_item, alignment=Qt.AlignCenter)

            # Add code
            code_item = QLabel(image_label.code)
            cell_layout.addWidget(code_item, alignment=Qt.AlignCenter)

            # Add quantity
            quantity_item = QLabel(str(image_label.quantity))
            cell_layout.addWidget(quantity_item, alignment=Qt.AlignCenter)

            # Add checkbox
            checkbox_item = QCheckBox()
            cell_layout.addWidget(checkbox_item, alignment=Qt.AlignCenter)

            # Set margins and spacing
            cell_layout.setContentsMargins(5, 5, 5, 5)  # Adjust margins as needed
            cell_layout.setSpacing(5)  # Adjust spacing between items as needed

            # Create a container widget for the cell
            cell_widget = QWidget()
            cell_widget.setLayout(cell_layout)

            # Add the cell to the grid
            self.grid_layout.addWidget(cell_widget, row, col)

        # Resize the grid layout
        self.grid_layout.setAlignment(Qt.AlignCenter)
        self.grid_layout.setContentsMargins(10, 10, 10, 10)  # Add margins to the grid

def main():
    """
    Entry point of the application.

    Initializes the QApplication and creates an instance of the 
    InvoiceSelectionWindow. This function sets up the main event loop 
    and displays the invoice selection window to the user.

    It uses the following steps:
    1. Creates a QApplication object with command line arguments.
    2. Instantiates the InvoiceSelectionWindow.
    3. Displays the window on the screen.
    4. Enters the application's main event loop until the user 
       closes the window.

    Exits the application cleanly upon closing the window.
    """
    app = QApplication(sys.argv)
    window = InvoiceSelectionWindow()
    window.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
